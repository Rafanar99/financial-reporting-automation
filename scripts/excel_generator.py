"""
Automated Excel Report Generator
Creates formatted Excel reports from consolidated financial data
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self, output_path='output/financial_report.xlsx'):
        self.output_path = output_path
        
    def create_summary_report(self, df, sheet_name='Summary'):
        df.to_excel(self.output_path, sheet_name=sheet_name, index=False)
        self._apply_formatting(sheet_name)
        print(f"Report generated: {self.output_path}")
        
    def _apply_formatting(self, sheet_name):
        wb = load_workbook(self.output_path)
        ws = wb[sheet_name]
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(self.output_path)
        
    def create_multi_sheet_report(self, data_dict):
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        for sheet_name in data_dict.keys():
            self._apply_formatting(sheet_name)
        print(f"Multi-sheet report generated: {self.output_path}")

if __name__ == "__main__":
    sample_data = pd.DataFrame({
        'Entity': ['Brazil', 'Argentina', 'Colombia'],
        'Revenue_USD': [1200000, 850000, 620000],
        'EBITDA_USD': [240000, 170000, 124000],
        'EBITDA_Margin': [20.0, 20.0, 20.0]
    })
    generator = ExcelReportGenerator('sample_output.xlsx')
    generator.create_summary_report(sample_data)
