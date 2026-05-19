"""
report_generator.py — Geração do Excel (openpyxl) e dos gráficos (matplotlib)
"""

import os
import io
import numpy as np
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.colors import LinearSegmentedColormap

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.drawing.image import Image as XLImage

matplotlib.use("Agg")  # backend sem GUI — essencial para ambientes sem display


# ──────────────────────────────────────────────
# Paleta de cores do relatório
# ──────────────────────────────────────────────

HEADER_BG   = "1F3864"   # azul escuro
HEADER_FONT = "FFFFFF"   # branco
ALT_ROW     = "EBF0FA"   # azul-claro para linhas alternadas
GREEN_FILL  = "C6EFCE"   # verde claro (retorno positivo)
RED_FILL    = "FFC7CE"   # vermelho claro (retorno negativo)
GREEN_FONT  = "276221"
RED_FONT    = "9C0006"


# ──────────────────────────────────────────────
# Helpers de estilo
# ──────────────────────────────────────────────

def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(cell, text: str) -> None:
    cell.value = text
    cell.font  = Font(bold=True, color=HEADER_FONT, name="Calibri", size=11)
    cell.fill  = PatternFill("solid", fgColor=HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _set_col_widths(ws, widths: dict) -> None:
    """widths = {col_letter: width_float}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _freeze(ws, cell: str) -> None:
    ws.freeze_panes = cell


# ──────────────────────────────────────────────
# Aba 1 — Raw Data
# ──────────────────────────────────────────────

def _sheet_raw_data(wb: Workbook, prices: pd.DataFrame) -> None:
    ws = wb.create_sheet("Raw Data")

    tickers = list(prices.columns)
    headers = ["Date"] + tickers

    # Cabeçalho
    for col_idx, h in enumerate(headers, start=1):
        _header_style(ws.cell(row=1, column=col_idx), h)

    ws.row_dimensions[1].height = 30

    # Dados
    for row_idx, (date, row) in enumerate(prices.iterrows(), start=2):
        # Coluna de data
        date_cell = ws.cell(row=row_idx, column=1, value=date.date())
        date_cell.number_format = "YYYY-MM-DD"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.border = _thin_border()

        fill_color = ALT_ROW if row_idx % 2 == 0 else "FFFFFF"

        for col_idx, ticker in enumerate(tickers, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=round(float(row[ticker]), 2))
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = _thin_border()

    # Larguras
    widths = {"A": 14}
    widths.update({get_column_letter(i + 2): 12 for i in range(len(tickers))})
    _set_col_widths(ws, widths)
    _freeze(ws, "B2")

    print("[Report] Aba 'Raw Data' criada. ✓")


# ──────────────────────────────────────────────
# Aba 2 — KPI Summary
# ──────────────────────────────────────────────

def _sheet_kpi_summary(wb: Workbook, kpis: pd.DataFrame) -> None:
    ws = wb.create_sheet("KPI Summary")

    col_headers = [
        "Ranking", "Ticker",
        "Preço Inicial (USD)", "Preço Final (USD)",
        "Retorno Acumulado (%)", "Volatilidade (% a.a.)",
        "Melhor Dia (%)", "Pior Dia (%)"
    ]

    # Cabeçalho
    for col_idx, h in enumerate(col_headers, start=1):
        _header_style(ws.cell(row=1, column=col_idx), h)
    ws.row_dimensions[1].height = 36

    # Dados
    for row_idx, row in kpis.iterrows():
        excel_row = row_idx + 2

        values = [
            row["Ranking"],
            row["Ticker"],
            round(row["Preço Inicial"], 2),
            round(row["Preço Final"], 2),
            round(row["Retorno Acumulado %"], 2),
            round(row["Volatilidade % a.a."], 2),
            round(row["Melhor Dia %"], 2),
            round(row["Pior Dia %"], 2),
        ]
        formats = [
            "0", "@",
            "#,##0.00", "#,##0.00",
            "0.00%", "0.00%",
            "0.00%", "0.00%",
        ]

        fill_color = ALT_ROW if excel_row % 2 == 0 else "FFFFFF"

        for col_idx, (val, fmt) in enumerate(zip(values, formats), start=1):
            # Percentuais já estão em %, precisamos dividir por 100 para o formato %
            if fmt == "0.00%" and isinstance(val, float):
                val = val / 100

            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center" if col_idx in (1, 2) else "right")
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = _thin_border()

        ws.row_dimensions[excel_row].height = 22

    # Formatação condicional — coluna "Retorno Acumulado %" (coluna E = 5)
    last_row = kpis.shape[0] + 1
    ret_col_range = f"E2:E{last_row}"

    ws.conditional_formatting.add(
        ret_col_range,
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["0"],
            fill=PatternFill("solid", fgColor=GREEN_FILL),
            font=Font(color=GREEN_FONT, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        ret_col_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=RED_FILL),
            font=Font(color=RED_FONT, bold=True),
        ),
    )

    _set_col_widths(ws, {
        "A": 10, "B": 10, "C": 20, "D": 18,
        "E": 22, "F": 22, "G": 16, "H": 14,
    })
    _freeze(ws, "C2")

    print("[Report] Aba 'KPI Summary' criada. ✓")


# ──────────────────────────────────────────────
# Gráfico 1 — Performance Chart (matplotlib)
# ──────────────────────────────────────────────

def _build_performance_chart(cumulative: pd.DataFrame, path: str) -> None:
    fig, ax = plt.subplots(figsize=(14, 7))
    fig.patch.set_facecolor("#F8F9FA")
    ax.set_facecolor("#F8F9FA")

    colors = plt.cm.tab10.colors

    for i, ticker in enumerate(cumulative.columns):
        ax.plot(
            cumulative.index,
            cumulative[ticker] * 100,
            label=ticker,
            linewidth=2,
            color=colors[i % len(colors)],
        )

    ax.axhline(0, color="black", linewidth=0.8, linestyle="--", alpha=0.5)
    ax.set_title("Retorno Acumulado — Últimos 12 Meses", fontsize=16, fontweight="bold", pad=15)
    ax.set_xlabel("Data", fontsize=12)
    ax.set_ylabel("Retorno Acumulado (%)", fontsize=12)
    ax.legend(loc="upper left", framealpha=0.9, fontsize=10)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}%"))

    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"[Report] Gráfico de performance salvo em '{path}'. ✓")


def _sheet_performance_chart(wb: Workbook, cumulative: pd.DataFrame, chart_path: str) -> None:
    ws = wb.create_sheet("Performance Chart")
    ws.sheet_view.showGridLines = False

    title_cell = ws["B2"]
    title_cell.value = "Retorno Acumulado — Últimos 12 Meses"
    title_cell.font = Font(bold=True, size=14, color=HEADER_BG)

    img = XLImage(chart_path)
    img.anchor = "B4"
    ws.add_image(img)

    print("[Report] Aba 'Performance Chart' criada. ✓")


# ──────────────────────────────────────────────
# Gráfico 2 — Correlation Heatmap (matplotlib)
# ──────────────────────────────────────────────

def _build_heatmap(correlation: pd.DataFrame, path: str) -> None:
    n = len(correlation)
    fig, ax = plt.subplots(figsize=(10, 8))
    fig.patch.set_facecolor("#F8F9FA")

    # Colormap: vermelho (-1) → branco (0) → verde escuro (+1)
    cmap = LinearSegmentedColormap.from_list(
        "corr_map",
        ["#C0392B", "#FFFFFF", "#1E8449"],
        N=256,
    )

    im = ax.imshow(correlation.values, cmap=cmap, vmin=-1, vmax=1, aspect="auto")

    # Rótulos dos eixos
    ax.set_xticks(range(n))
    ax.set_yticks(range(n))
    ax.set_xticklabels(correlation.columns, rotation=45, ha="right", fontsize=11)
    ax.set_yticklabels(correlation.index, fontsize=11)

    # Valores dentro de cada célula
    for i in range(n):
        for j in range(n):
            val = correlation.iloc[i, j]
            text_color = "white" if abs(val) > 0.6 else "black"
            ax.text(j, i, f"{val:.2f}", ha="center", va="center",
                    fontsize=10, color=text_color, fontweight="bold")

    plt.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="Correlação de Pearson")
    ax.set_title("Matriz de Correlação entre Ativos", fontsize=15, fontweight="bold", pad=12)

    plt.tight_layout()
    plt.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"[Report] Heatmap de correlação salvo em '{path}'. ✓")


def _sheet_correlation_heatmap(wb: Workbook, heatmap_path: str) -> None:
    ws = wb.create_sheet("Correlation Heatmap")
    ws.sheet_view.showGridLines = False

    title_cell = ws["B2"]
    title_cell.value = "Matriz de Correlação entre Ativos"
    title_cell.font = Font(bold=True, size=14, color=HEADER_BG)

    img = XLImage(heatmap_path)
    img.anchor = "B4"
    ws.add_image(img)

    print("[Report] Aba 'Correlation Heatmap' criada. ✓")


# ──────────────────────────────────────────────
# CSV para Power BI
# ──────────────────────────────────────────────

def save_csv(kpis: pd.DataFrame, daily_returns: pd.DataFrame, csv_path: str) -> None:
    """
    Combina KPIs e retornos diários em um único CSV,
    formato long (tidy) — ideal para o Power BI.
    """
    # Retornos diários no formato long
    daily_long = (
        daily_returns
        .reset_index()
        .melt(id_vars=daily_returns.index.name or "Date", var_name="Ticker", value_name="Retorno Diário")
    )
    daily_long.rename(columns={daily_long.columns[0]: "Date"}, inplace=True)
    daily_long["Retorno Diário %"] = daily_long["Retorno Diário"] * 100

    # Merge com KPIs
    kpi_cols = ["Ticker", "Preço Inicial", "Preço Final",
                "Retorno Acumulado %", "Volatilidade % a.a.",
                "Melhor Dia %", "Pior Dia %", "Ranking"]
    merged = daily_long.merge(kpis[kpi_cols], on="Ticker", how="left")

    merged.to_csv(csv_path, index=False, float_format="%.4f")
    print(f"[Report] CSV salvo em '{csv_path}'. ✓")


# ──────────────────────────────────────────────
# Função principal
# ──────────────────────────────────────────────

def generate_report(
    data: dict,
    output_dir: str = "output",
    assets_dir: str = "assets",
) -> None:
    """
    Recebe o dicionário produzido por processor.process_all() e gera:
        - output/financial_report.xlsx
        - output/portfolio_data.csv
        - assets/performance_chart.png
        - assets/correlation_heatmap.png
    """
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(assets_dir, exist_ok=True)

    chart_path  = os.path.join(assets_dir, "performance_chart.png")
    heatmap_path = os.path.join(assets_dir, "correlation_heatmap.png")
    excel_path  = os.path.join(output_dir, "financial_report.xlsx")
    csv_path    = os.path.join(output_dir, "portfolio_data.csv")

    prices      = data["prices"]
    daily       = data["daily"]
    cumulative  = data["cumulative"]
    kpis        = data["kpis"]
    correlation = data["correlation"]

    # ── Gráficos ──
    print("\n[Report] Gerando gráficos...")
    _build_performance_chart(cumulative, chart_path)
    _build_heatmap(correlation, heatmap_path)

    # ── Workbook ──
    print("\n[Report] Construindo workbook Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # remove a aba padrão vazia

    _sheet_raw_data(wb, prices)
    _sheet_kpi_summary(wb, kpis)
    _sheet_performance_chart(wb, cumulative, chart_path)
    _sheet_correlation_heatmap(wb, heatmap_path)

    wb.save(excel_path)
    print(f"\n[Report] Excel salvo em '{excel_path}'. ✓")

    # ── CSV ──
    save_csv(kpis, daily, csv_path)

    print("\n✅  Relatório gerado com sucesso!")
    print(f"   Excel  → {excel_path}")
    print(f"   CSV    → {csv_path}")
    print(f"   Charts → {assets_dir}/")
